
import openpyxl
from openpyxl.styles import PatternFill

def compare_excel_files(file_a, file_b,output_file):
    # Load the Excel files
    wb_a = openpyxl.load_workbook(file_a)
    wb_b = openpyxl.load_workbook(file_b)
    
    # Get the active sheets
    sheet_a = wb_a.active
    sheet_b = wb_b.active

    # Create red fill style
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    
    # Initialize record, wrong_number, and pos lists
    record =[]
    wrong_number = []
    pos =[]
    
    # Iterate through columns in the range 1 to 10 (inclusive)
    for column in range(1, 11):

        column_a = list(list(sheet_a.iter_cols(min_col=column, max_col=column, values_only=True))[0])
        column_b = list(list(sheet_b.iter_cols(min_col=column, max_col=column, values_only=True))[0])
        
        # Initialize store, cur, k, and count variables
        store = []
        cur = 0
        k = 0
        count = 0
        
        # Compare elements in column_a and column_b
        while k < len(column_b):
            
            #handle the column 'Dimensions'
            if column == 10 and column_a[k+count+cur]!= None and column_b[k]!= None and column_a[k+count+cur].split(':')[0] == column_b[k].split(':')[0]:
                
                wrong = []
                if column_a[k+count+cur]!=(column_b[k]):
                    store.append(k+count+cur)
                    wrong_number.append(k+count+cur)
                    for l in range(1,6):
                        cell_a =sheet_a.cell(row = k+count+cur+1, column = column +l).value
                        cell_b = sheet_b.cell(row = k+1, column = column +l).value
                        if cell_a != cell_b:
                            wrong.append(l)
                    pos.append(wrong)
                    
            #handle other columns
            else:
                if column_a[k+count]!=(column_b[k]):
                    cur = cur + count
                    count =0
                while column_a[k+count+cur]!=(column_b[k]):
                    
                    if column_a[k+count+cur] != None:
                        store.append(k+count+cur)
                    count += 1
            k = k+1
            if column_b[k-1] != None:
                last_word = column_b[k-1]
        
        # Check for additional elements in column_a
        if len(column_a) >len(column_b):
            diff = len(column_a) -column_a.index(last_word)
            for j in range(1,diff):

                if column_a[j+column_a.index(last_word)] != None:
                    store.append(j+column_a.index(last_word))
            record.append(store)
    count = 0
    
    #Fill the cell (with incomplete or wrong file name) with red color
    for i in range(len(record)):
        for j in range(len(record[i])):
            
            if i == len(record)-1 and (record[i][j] not in wrong_number):
                for k in range(0,6):
                    cell =  sheet_a.cell(row = record[i][j]+1, column = i+1+k)
                    if cell != None:
                        cell.fill = red_fill
            elif i == len(record)-1 and (record[i][j] in wrong_number):
                cell =  sheet_a.cell(row = record[i][j]+1, column = i+1)
                cell.fill = red_fill
                for c in pos[count]:
                    cell = sheet_a.cell(row = record[i][j]+1, column = i+1+c)
                    cell.fill = red_fill
                count += 1
                
            else:
                cell = sheet_a.cell(row = record[i][j]+1, column = i+1)
                cell.fill = red_fill


    wb_a.save(output_file)

if __name__ == '__main__':
    file_a = 'CMRxRecon.xlsx'
    file_b = 'CMRxRecon_check.xlsx' 
    output_file = 'Comparison_Result.xlsx'  
    
    compare_excel_files(file_a, file_b, output_file)