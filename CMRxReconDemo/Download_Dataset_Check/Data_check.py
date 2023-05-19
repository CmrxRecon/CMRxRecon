import os
import openpyxl
import mat73

def create_directory_structure_excel(directory_path, output_file):
    # Create a new workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Recursive function to process subdirectories and files
    def process_directory(directory, level, folder_name=None):
        nonlocal sheet

        # Iterate over subdirectories and files
        for name in sorted(os.listdir(directory)):
            path = os.path.join(directory, name)

            if os.path.isdir(path):
                # Write the directory name to the Excel sheet
                sheet.cell(row=len(sheet['A']) + 1, column=level + 1).value = name
                process_directory(path, level + 1, name)
            else:
                # Write the file name to the Excel sheet
                sheet.cell(row=len(sheet['A']) + 1, column=level + 1).value = str(path)
                if '.mat' in name:
                    # Read .mat file
                    data = mat73.loadmat(path)
                    variable_name = list(data.keys())[0]

                    # Get the dimensions of the matrix
                    matrix = data[variable_name]
                    dimensions = matrix.shape

                    # Combine folder name, colon, and dimensions
                    if folder_name is not None:
                        folder_info = f"{folder_name}/{name}: {str(dimensions)}"
                    else:
                        folder_info = str(dimensions)

                    sheet.cell(row=len(sheet['A']), column=level + 3).value = folder_info

                    # Split dimensions into individual numbers
                    dimensions_list = list(dimensions)
                    row_index = len(sheet['A'])

                    # Write each dimension number to subsequent cells
                    for i, dimension in enumerate(dimensions_list):
                        sheet.cell(row=row_index, column=level + 4 + i).value = dimension

    # Start processing the directory structure
    process_directory(directory_path, 0)

    # Add column headers for Level 1, Level 2, etc.
    num_columns = 9
    for column in range(1, num_columns + 1):
        if column == num_columns:
            sheet.cell(row=1, column=column).value = "Dimensions"
        else:
            sheet.cell(row=1, column=column).value = f"Level {column}"

    # Save the workbook to the output file
    workbook.save(output_file)

if __name__ == '__main__':
    # directory_path = "CMRxRecon"  # Replace with the actual directory path
    # output_file = "CMRxRecon_check.xlsx"  # Specify the output file name
    directory_path='/Users/xinyuzhang/Desktop/hi'
    output_file = '/Users/xinyuzhang/Desktop/test.xlsx'

    create_directory_structure_excel(directory_path, output_file)